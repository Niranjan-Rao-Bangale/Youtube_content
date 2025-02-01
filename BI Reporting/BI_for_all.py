import json
import re
import pandas as pd
import streamlit as st
import google.generativeai as genai
import plotly.express as px  
import io 
from io import BytesIO
import base64  
from sqlalchemy import create_engine, text
from fpdf import FPDF
import openpyxl

# ----------------- DATABASE CONFIGURATION -------------------
db_path = "<Your Database connection>/DBeaverData/workspace6/.metadata/sample-database-sqlite-1/Chinook.db" # you use sql lite database or it can be any database for example you can also use postgresql database with this syntax: "postgresql+psycopg2://<User>:<Password>@<Host>:<port>/<database> 
engine = create_engine(db_path)

# ----------------- CONFIGURE AI MODEL ----------------------
genai.configure(api_key='<Your Gemini API Key>')
model = genai.GenerativeModel('gemini-1.5-flash')

# ----------------- STREAMLIT PAGE CONFIG -------------------
st.set_page_config(page_title="BI Report Generator", page_icon="📊", layout="wide")

# ----------------- SESSION STATE INITIALIZATION -------------------
if "sql_query" not in st.session_state:
    st.session_state.sql_query = None
if "query_results" not in st.session_state:
    st.session_state.query_results = None
if "query_columns" not in st.session_state:
    st.session_state.query_columns = None

# ----------------- STYLIZED PAGE HEADER -------------------
st.markdown(
    """
    <style>
        .big-title {
            font-size: 30px !important;
            font-weight: bold;
            text-align: center;
        }
        .export-container {
            padding: 15px;
            border-radius: 10px;
            background-color: #f8f9fa;
            margin-top: 10px;
        }
        .stButton>button {
            width: 100%;
            border-radius: 10px;
            padding: 10px;
        }
    </style>
    """,
    unsafe_allow_html=True
)

st.markdown('<p class="big-title">📊 BI Report Generator</p>', unsafe_allow_html=True)


# ----------------- AI SQL GENERATION FUNCTION -------------------
# You need to pass your databse table schema information so that LLM models can understand table structure and generate queries accondingly.
def generate_sql_with_llm(prompt):
    full_prompt = f"""
    You are an AI assistant that generates SQL queries based on the user's request. 
    It should follow sql lite database syntax.
    The queries must match the following database schema (SQLite):
    1. album (AlbumId, Title, ArtistId)
    2. artist (artistid, name)
    3. track (trackid, name, albumid, genreid, composer, milliseconds, bytes, unit_price, MediaTypeId)
    4. Invoice (InvoiceId, CustomerId, InvoiceDate, BillingAddress, BillingCity, BillingState, BillingCountry, 
    BillingPostalCode, Total)
    5. InvoiceLine (InvoiceLineId, InvoiceId, TrackId, UnitPrice, Quantity)
    6. MediaType (MediaTypeId, Name)
    7. Genre (GenreId, Name)
    8. Customer (CustomerId, FirstName, LastName, Company, Address, City, State, Country, PostalCode, Phone,
    Fax, Email, SupportRepId)
    9. Employee (EmployeeId, LastName, FirstName, Title, ReportsTo, BirthDate, HireDate, Address, City, State,
    Country, PostalCode, Phone, Fax, Email)
    10. Playlist (PlaylistId, Name)
    11. PlaylistTrack (PlaylistId, TrackId)

    Convert the user's natural language request into a valid SQL query:
    Request: "{prompt}"
    Output JSON format: {{"query": "<SQL query>"}}
    """

    model_response = model.generate_content(full_prompt)
    response_text = model_response.text.strip()
    if response_text.startswith("```json") and response_text.endswith("```"):
        response_text = response_text[7:-3].strip()

    try:
        data = json.loads(response_text)
        return data.get("query", None)
    except json.JSONDecodeError:
        return None

# ----------------- SQL EXECUTION FUNCTION -------------------
def execute_sql_query(query):
    try:
        with engine.connect() as connection:
            result = connection.execute(text(query))
            df = pd.DataFrame(result.fetchall(), columns=result.keys())
        return df
    except Exception as e:
        st.error(f"Error executing query: {e}")
        return pd.DataFrame()

# ----------------- CHART GENERATION FUNCTION -------------------
def plot_interactive_chart(df, chart_type):
    x_column = df.columns[0]
    numeric_columns = df.select_dtypes(include=["number"]).columns.tolist()

    if not numeric_columns:
        st.warning("No numeric columns available for visualization.")
        return

    y_columns = [col for col in numeric_columns if col != x_column]
    if not y_columns:
        st.warning("No valid numeric data to plot.")
        return

    if chart_type == "Bar Chart":
        fig = px.bar(df, x=x_column, y=y_columns, title="Interactive Bar Chart")
    elif chart_type == "Line Chart":
        fig = px.line(df, x=x_column, y=y_columns, title="Interactive Line Chart")
    elif chart_type == "Pie Chart" and len(y_columns) == 1:
        fig = px.pie(df, names=x_column, values=y_columns[0], title="Interactive Pie Chart")
    else:
        st.warning("Pie Chart requires exactly one numeric column.")
        return

    st.plotly_chart(fig, use_container_width=True)

# ----------------- EXPORT FUNCTIONS -------------------
def generate_csv(df):
    return df.to_csv(index=False).encode("utf-8")

def generate_excel(df):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine="openpyxl")
    df.to_excel(writer, sheet_name="Report", index=False)
    writer.close()
    output.seek(0)
    return output


def generate_pdf(df):
    """
    Generate a PDF from the given DataFrame and return its bytes.
    """
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    # Title
    pdf.set_font("Arial", style='B', size=14)
    pdf.cell(200, 10, "Report", ln=True, align="C")
    pdf.ln(10)

    # Set column headers
    pdf.set_font("Arial", style='B', size=10)
    col_width = 40  # Fixed column width
    for col_name in df.columns:
        pdf.cell(col_width, 10, col_name, border=1)
    pdf.ln()

    # Add rows
    pdf.set_font("Arial", size=10)
    for _, row in df.iterrows():
        for value in row:
            pdf.cell(col_width, 10, str(value), border=1)
        pdf.ln()

    # Return PDF as bytes:
    pdf_string = pdf.output(dest="S")  # Returns a string
    pdf_bytes = pdf_string.encode("latin-1")  # Convert to bytes
    return pdf_bytes



# ----------------- SIDEBAR FOR QUERY INPUT -------------------
st.sidebar.title("🔍 Search & Filter Data")
user_input = st.sidebar.text_area("Enter your query:", "Show me the total sales per country")

if st.sidebar.button("Generate & Run Query"):
    sql_query = generate_sql_with_llm(user_input)

    if sql_query:
        st.session_state.sql_query = sql_query
        st.sidebar.code(sql_query, language="sql")

        df = execute_sql_query(sql_query)
        if df.empty:
            st.error("No data found.")
        else:
            st.session_state.query_results = df
    else:
        st.warning("Could not generate a valid SQL query. Please try again.")

# ----------------- DISPLAY RESULTS & EXPORT OPTIONS -------------------
if st.session_state.query_results is not None:
    st.markdown("<a name='results_section'></a>", unsafe_allow_html=True)

    df = st.session_state.query_results

    chart_type = st.selectbox("📊 Select Chart Type", ["Table", "Bar Chart", "Line Chart", "Pie Chart"])

    if chart_type == "Table":
        st.write(df)
    else:
        plot_interactive_chart(df, chart_type)

    with st.expander("📤 Export Report"):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.download_button("📄 Download PDF", data=generate_pdf(df), file_name="report.pdf", mime="application/pdf")
        with col2:
            st.download_button("📊 Download Excel", data=generate_excel(df), file_name="report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with col3:
            st.download_button("📜 Download CSV", data=generate_csv(df), file_name="report.csv", mime="text/csv")

    # 2) Inject JavaScript to scroll to the anchor AFTER displaying the results
    scroll_script = """
    <script>
        document.getElementsByName('results_section')[0].scrollIntoView({behavior: 'smooth'});
    </script>
    """
    st.markdown(scroll_script, unsafe_allow_html=True)
